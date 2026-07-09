"""Monitor monthly GB and industry standard announcements."""

from __future__ import annotations

import argparse
import json
import re
import smtplib
import time
import urllib.parse
import urllib.request
from dataclasses import asdict, dataclass
from datetime import date, datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from html import unescape
from pathlib import Path
from typing import Any

USER_AGENT = "Mozilla/5.0 (compatible; Standard-Notice-Monitor/1.2)"
REQUEST_DELAY = 1.5

DOMAIN_KEYWORDS: dict[str, list[str]] = {
    "化工": ["化工", "化学", "石化", "石油", "塑料", "橡胶", "涂料", "胶粘", "树脂", "溶剂", "化肥", "农药"],
    "新材料": ["新材料", "复合材料", "纳米", "碳纤维", "石墨烯", "稀土", "合金", "陶瓷材料", "功能材料", "超材料"],
    "电子材料": ["电子材料", "半导体", "介电", "磁性材料", "封装材料", "压电", "光电", "靶材", "溅射", "晶圆材料"],
    "集成电路": ["集成电路", "芯片", "晶圆", "微电子", "光刻", "EDA", "封装测试", "IC", "存储器", "处理器"],
    "安全": [
        "安全",
        "安全生产",
        "职业安全",
        "职业健康",
        "防爆",
        "危险化学品",
        "重大危险源",
        "应急",
        "压力容器",
        "特种设备",
        "锅炉",
        "气瓶",
    ],
    "消防": ["消防", "防火", "灭火", "火灾", "消防救援", "消防设施", "阻燃", "消防器材"],
    "生产": ["生产", "生产工艺", "生产过程", "制造", "工业化生产", "清洁生产"],
    "合成生物": ["合成生物", "合成生物学", "生物制造", "基因工程", "细胞工厂", "发酵", "生物基"],
    "产业园区": ["产业园区", "工业园", "化工园区", "开发区", "产业集聚"],
    "中试": ["中试", "中间试验", "试验规模", "小试"],
    "中试基地": ["中试基地", "中试平台", "中试线", "中试车间"],
    "科技创新": ["科技创新", "技术创新", "科技成果转化", "研发", "自主创新", "科技攻关"],
}

TRACKED_INDUSTRY_CODES = ("AQ", "GA", "HG", "HJ")
INDUSTRY_CODE_PATTERN = re.compile(r"^(AQ|GA|HG|HJ)(?:/T|\s)", re.I)

SHANGHAI_NOTICE_KEY = "上海"
SHANGHAI_STD_PATTERN = re.compile(r"^DB31(?:/T|\s)", re.I)
DBBA_BASE = "https://dbba.sacinfo.org.cn"
HBBA_BASE = "https://hbba.sacinfo.org.cn"

STD_ROW_PATTERN = re.compile(
    r"<tr>\s*<td>(\d+)</td>\s*<td[^>]*>([^<]+)</td>\s*<td>.*?data-param=\"(\{[^\"]+\})\".*?>"
    r"([^<]+)</a>.*?</tr>",
    re.DOTALL,
)
SACINFO_TAG_PATTERN = re.compile(r"<sacinfo>([^<]*)</sacinfo>")


@dataclass
class StandardItem:
    std_type: str
    industry_code: str
    notice_code: str
    notice_date: str
    notice_title: str
    notice_url: str
    std_code: str
    std_name: str
    implement_date: str
    matched_domains: str
    info_url: str
    download_url: str


def standard_item_to_dict(item: StandardItem) -> dict[str, Any]:
    return asdict(item)


def collect_all(
    year: int,
    month: int,
    *,
    run_gb: bool = True,
    run_hb: bool = True,
    run_db: bool = True,
) -> tuple[list[StandardItem], list[StandardItem], list[StandardItem]]:
    gb_items = collect_gb_items(year, month) if run_gb else []
    hb_items = collect_hb_items(year, month) if run_hb else []
    db_items = collect_db_items(year, month) if run_db else []
    return gb_items, hb_items, db_items


def _request(url: str, *, params: dict[str, Any] | None = None, method: str = "GET", data: bytes | None = None) -> bytes:
    if params and method == "GET":
        query = urllib.parse.urlencode(params)
        sep = "&" if "?" in url else "?"
        url = f"{url}{sep}{query}"
    headers = {"User-Agent": USER_AGENT}
    if data is not None:
        headers["Content-Type"] = "application/x-www-form-urlencoded"
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


def _post_form(url: str, form: dict[str, Any]) -> dict[str, Any]:
    body = _request(url, method="POST", data=urllib.parse.urlencode(form).encode("utf-8"))
    return json.loads(body.decode("utf-8"))


def _sleep() -> None:
    time.sleep(REQUEST_DELAY)


def _clean(text: str) -> str:
    text = unescape(re.sub(r"<[^>]+>", " ", text or ""))
    return re.sub(r"\s+", " ", text).strip()


def _ts_to_date(ms: int | float | None) -> str:
    if not ms:
        return ""
    return datetime.fromtimestamp(ms / 1000).strftime("%Y-%m-%d")


def _industry_code(std_code: str) -> str | None:
    match = INDUSTRY_CODE_PATTERN.match(std_code.strip())
    return match.group(1).upper() if match else None


def _is_shanghai_std(std_code: str, std: dict[str, Any] | None = None) -> bool:
    if SHANGHAI_STD_PATTERN.match(std_code.strip()):
        return True
    if std:
        for field in ("chargeDept", "industry", "recordDept"):
            if "上海" in (std.get(field) or ""):
                return True
    return False


def _fetch_sacinfo_notices_for_month(
    base_url: str,
    year: int,
    month: int,
    *,
    keyword: str = "",
) -> list[dict[str, Any]]:
    target_prefix = f"{year:04d}-{month:02d}"
    page = 1
    page_size = 50
    matched: list[dict[str, Any]] = []

    while True:
        payload = _post_form(
            f"{base_url}/snQueryList",
            {"current": page, "size": page_size, "key": keyword, "snType": "pub"},
        )
        _sleep()
        rows = payload.get("records") or []
        if not rows:
            break

        for row in rows:
            release_date = _ts_to_date(row.get("releaseDate"))
            if release_date.startswith(target_prefix):
                row["releaseDateStr"] = release_date
                matched.append(row)

        oldest = min((_ts_to_date(r.get("releaseDate")) for r in rows), default="")
        if oldest and oldest < target_prefix:
            break
        if page >= int(payload.get("pages", 0)):
            break
        page += 1

    return matched


def _fetch_sacinfo_notice_standards(base_url: str, notice_pk: str) -> list[dict[str, Any]]:
    standards: list[dict[str, Any]] = []
    page = 1
    page_size = 200

    while True:
        payload = _post_form(
            f"{base_url}/stdQueryList?snfzId={notice_pk}",
            {"current": page, "size": page_size},
        )
        _sleep()
        rows = payload.get("records") or []
        standards.extend(rows)
        if not rows or page >= int(payload.get("pages", 0)):
            break
        page += 1

    return standards


# --- 国家标准 (std.samr.gov.cn) ---


def fetch_gb_notices_for_month(year: int, month: int) -> list[dict[str, str]]:
    target_prefix = f"{year:04d}-{month:02d}"
    page = 1
    page_size = 50
    matched: list[dict[str, str]] = []

    while True:
        body = _request(
            "https://std.samr.gov.cn/noc/search/nocGBPage",
            params={"searchText": "", "pageNumber": page, "pageSize": page_size},
        )
        _sleep()
        payload = json.loads(body.decode("utf-8"))
        rows = payload.get("rows") or []
        if not rows:
            break

        for row in rows:
            if row.get("NOTICE_DATE", "").startswith(target_prefix):
                matched.append(row)

        oldest = min((r.get("NOTICE_DATE", "") for r in rows), default="")
        if oldest and oldest < target_prefix:
            break
        if page * page_size >= int(payload.get("total", 0)):
            break
        page += 1

    return matched


def parse_gb_notice_standards(pid: str) -> tuple[str, list[dict[str, str]]]:
    url = f"http://std.sacinfo.org.cn/gnoc/queryInfo?id={pid}"
    html = _request(url).decode("utf-8", "replace")
    _sleep()

    title_match = re.search(r'<p class="ntitle">(.*?)</p>', html, re.DOTALL)
    notice_title = _clean(title_match.group(1)) if title_match else ""

    standards: list[dict[str, str]] = []
    for _, std_code, _param, std_name in STD_ROW_PATTERN.findall(html):
        standards.append({"std_code": _clean(std_code), "std_name": _clean(std_name)})
    return notice_title, standards


def lookup_gb_standard_links(std_code: str) -> tuple[str, str]:
    body = _request(
        "https://std.samr.gov.cn/gb/search/gbQueryPage",
        params={"searchText": std_code, "pageNumber": 1, "pageSize": 5},
    )
    _sleep()
    payload = json.loads(body.decode("utf-8"))
    rows = payload.get("rows") or []
    if not rows:
        return "", ""

    row = rows[0]
    hcno = row.get("id", "")
    code_parts = SACINFO_TAG_PATTERN.findall(row.get("C_STD_CODE", ""))
    normalized = " ".join(code_parts).strip() if code_parts else std_code

    if normalized.replace(" ", "") != std_code.replace(" ", ""):
        for candidate in rows:
            parts = SACINFO_TAG_PATTERN.findall(candidate.get("C_STD_CODE", ""))
            cand_code = " ".join(parts).strip()
            if cand_code.replace(" ", "") == std_code.replace(" ", ""):
                hcno = candidate.get("id", "")
                break

    info_url = f"https://openstd.samr.gov.cn/bzgk/gb/newGbInfo?hcno={hcno}"
    download_url = f"https://openstd.samr.gov.cn/bzgk/std/showGb?type=download&hcno={hcno}"
    return info_url, download_url


def match_domains(std_name: str, std_code: str) -> list[str]:
    text = f"{std_name} {std_code}"
    return [domain for domain, keywords in DOMAIN_KEYWORDS.items() if any(kw in text for kw in keywords)]


def collect_gb_items(year: int, month: int) -> list[StandardItem]:
    items: list[StandardItem] = []
    for notice in fetch_gb_notices_for_month(year, month):
        pid = notice["PID"]
        notice_code = notice.get("CODE", "")
        notice_date = notice.get("NOTICE_DATE", "")
        notice_url = f"http://std.sacinfo.org.cn/gnoc/queryInfo?id={pid}"

        notice_title, standards = parse_gb_notice_standards(pid)
        if not notice_title:
            notice_title = notice.get("TITLE", "")

        for std in standards:
            domains = match_domains(std["std_name"], std["std_code"])
            if not domains:
                continue
            info_url, download_url = lookup_gb_standard_links(std["std_code"])
            items.append(
                StandardItem(
                    std_type="国标",
                    industry_code="",
                    notice_code=notice_code,
                    notice_date=notice_date,
                    notice_title=notice_title,
                    notice_url=notice_url,
                    std_code=std["std_code"],
                    std_name=std["std_name"],
                    implement_date="",
                    matched_domains="、".join(domains),
                    info_url=info_url,
                    download_url=download_url,
                )
            )
    return items


# --- 行业标准 (hbba.sacinfo.org.cn) ---


def fetch_hb_notices_for_month(year: int, month: int) -> list[dict[str, Any]]:
    return _fetch_sacinfo_notices_for_month(HBBA_BASE, year, month)


def fetch_hb_notice_standards(notice_pk: str) -> list[dict[str, Any]]:
    return _fetch_sacinfo_notice_standards(HBBA_BASE, notice_pk)


def collect_hb_items(year: int, month: int) -> list[StandardItem]:
    items: list[StandardItem] = []
    for notice in fetch_hb_notices_for_month(year, month):
        notice_pk = notice["pk"]
        notice_code = notice.get("code", "")
        notice_date = notice.get("releaseDateStr", "")
        notice_title = notice_code
        notice_url = f"{HBBA_BASE}/snDetail/{notice_pk}"
        publisher = notice.get("createRealname", "")

        for std in fetch_hb_notice_standards(notice_pk):
            std_code = _clean(std.get("code", ""))
            industry = _industry_code(std_code)
            if industry not in TRACKED_INDUSTRY_CODES:
                continue

            std_pk = std.get("pk", "")
            items.append(
                StandardItem(
                    std_type="行标",
                    industry_code=industry,
                    notice_code=notice_code,
                    notice_date=notice_date,
                    notice_title=f"{notice_title}（发布单位：{publisher}）",
                    notice_url=notice_url,
                    std_code=std_code,
                    std_name=_clean(std.get("chName", "")),
                    implement_date=_ts_to_date(std.get("actDate")),
                    matched_domains=industry,
                    info_url=f"{HBBA_BASE}/stdDetail/{std_pk}",
                    download_url=f"{HBBA_BASE}/portal/online/{std_pk}",
                )
            )
    return items


# --- 地方标准-上海 (dbba.sacinfo.org.cn) ---


def collect_db_items(year: int, month: int) -> list[StandardItem]:
    items: list[StandardItem] = []
    for notice in _fetch_sacinfo_notices_for_month(DBBA_BASE, year, month, keyword=SHANGHAI_NOTICE_KEY):
        notice_pk = notice["pk"]
        notice_code = notice.get("code", "")
        notice_date = notice.get("releaseDateStr", "")
        notice_title = notice_code
        notice_url = f"{DBBA_BASE}/snDetail/{notice_pk}"
        publisher = notice.get("createRealname", "")

        for std in _fetch_sacinfo_notice_standards(DBBA_BASE, notice_pk):
            std_code = _clean(std.get("code", ""))
            if not _is_shanghai_std(std_code, std):
                continue

            std_pk = std.get("pk", "")
            items.append(
                StandardItem(
                    std_type="地标",
                    industry_code="上海",
                    notice_code=notice_code,
                    notice_date=notice_date,
                    notice_title=f"{notice_title}（发布单位：{publisher}）",
                    notice_url=notice_url,
                    std_code=std_code,
                    std_name=_clean(std.get("chName", "")),
                    implement_date=_ts_to_date(std.get("actDate")),
                    matched_domains="上海",
                    info_url=f"{DBBA_BASE}/stdDetail/{std_pk}",
                    download_url=f"{DBBA_BASE}/portal/online/{std_pk}",
                )
            )
    return items


# --- 导出与邮件 ---

EXCEL_HEADERS = [
    "标准类型",
    "行业代码",
    "公告号",
    "公告日期",
    "公告标题",
    "公告链接",
    "标准号",
    "标准名称",
    "实施日期",
    "匹配领域",
    "标准详情页",
    "下载/全文地址",
]


def _item_to_row(item: StandardItem) -> list[str]:
    return [
        item.std_type,
        item.industry_code,
        item.notice_code,
        item.notice_date,
        item.notice_title,
        item.notice_url,
        item.std_code,
        item.std_name,
        item.implement_date,
        item.matched_domains,
        item.info_url,
        item.download_url,
    ]


def export_excel(
    gb_items: list[StandardItem],
    hb_items: list[StandardItem],
    db_items: list[StandardItem],
    output: Path,
) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise SystemExit("请先安装 openpyxl: pip install openpyxl") from exc

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "汇总"
    ws_all.append(EXCEL_HEADERS)
    for item in gb_items + hb_items + db_items:
        ws_all.append(_item_to_row(item))

    ws_gb = wb.create_sheet("国标-领域筛选")
    ws_gb.append(EXCEL_HEADERS)
    for item in gb_items:
        ws_gb.append(_item_to_row(item))

    ws_hb = wb.create_sheet("行标-AQ_GA_HG_HJ")
    ws_hb.append(EXCEL_HEADERS)
    for item in hb_items:
        ws_hb.append(_item_to_row(item))

    ws_db = wb.create_sheet("地标-上海")
    ws_db.append(EXCEL_HEADERS)
    for item in db_items:
        ws_db.append(_item_to_row(item))

    wb.save(output)
    return output


def send_email(
    *,
    smtp_host: str,
    smtp_port: int,
    smtp_user: str,
    smtp_password: str,
    mail_to: str,
    subject: str,
    body: str,
    attachment: Path | None = None,
) -> None:
    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = mail_to
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    if attachment and attachment.exists():
        with attachment.open("rb") as f:
            part = MIMEApplication(f.read(), Name=attachment.name)
        part["Content-Disposition"] = f'attachment; filename="{attachment.name}"'
        msg.attach(part)

    with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def main() -> None:
    parser = argparse.ArgumentParser(description="爬取当月国标/行标/上海地标公告并筛选目标标准")
    parser.add_argument("--year", type=int, default=date.today().year)
    parser.add_argument("--month", type=int, default=date.today().month)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("output") / f"std_notice_{date.today():%Y%m}.xlsx",
    )
    parser.add_argument("--config", type=Path, default=Path("config.json"))
    parser.add_argument("--no-email", action="store_true", help="只生成表格，不发送邮件")
    parser.add_argument("--gb-only", action="store_true", help="仅抓取国标")
    parser.add_argument("--hb-only", action="store_true", help="仅抓取行业标准")
    parser.add_argument("--db-only", action="store_true", help="仅抓取上海市地方标准")
    args = parser.parse_args()

    only_flags = [args.gb_only, args.hb_only, args.db_only]
    if sum(only_flags) > 1:
        parser.error("--gb-only、--hb-only、--db-only 不能同时使用")
    run_gb = args.gb_only or not any(only_flags)
    run_hb = args.hb_only or not any(only_flags)
    run_db = args.db_only or not any(only_flags)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    period = f"{args.year}-{args.month:02d}"

    gb_items: list[StandardItem] = []
    hb_items: list[StandardItem] = []
    db_items: list[StandardItem] = []

    if run_gb:
        print(f"正在抓取 {period} 的国家标准公告...")
        gb_items = collect_gb_items(args.year, args.month)
        print(f"  国标领域匹配: {len(gb_items)} 条")

    if run_hb:
        print(f"正在抓取 {period} 的行业标准公告 (AQ/GA/HG/HJ)...")
        hb_items = collect_hb_items(args.year, args.month)
        print(f"  行标行业匹配: {len(hb_items)} 条")

    if run_db:
        print(f"正在抓取 {period} 的上海市地方标准公告...")
        db_items = collect_db_items(args.year, args.month)
        print(f"  上海地标: {len(db_items)} 条")

    export_excel(gb_items, hb_items, db_items, args.output)
    print(f"已保存: {args.output.resolve()}")

    if args.no_email:
        return

    cfg = load_config(args.config)
    email_cfg = cfg.get("email") or {}
    required = ["smtp_host", "smtp_port", "smtp_user", "smtp_password", "mail_to"]
    if not all(email_cfg.get(k) for k in required):
        print("未配置完整邮件信息，跳过发送。请复制 config.example.json 为 config.json 并填写邮箱。")
        return

    subject = f"标准公告筛选报告 {period}"
    body_lines = [
        f"报告生成时间: {datetime.now():%Y-%m-%d %H:%M:%S}",
        f"目标月份: {period}",
        f"国标（领域关键词筛选）: {len(gb_items)} 条",
        f"行标（AQ/GA/HG/HJ）: {len(hb_items)} 条",
        f"地标（上海市）: {len(db_items)} 条",
        "",
        "详见附件 Excel（含汇总、国标、行标、地标四个工作表）。",
    ]
    send_email(
        smtp_host=email_cfg["smtp_host"],
        smtp_port=int(email_cfg["smtp_port"]),
        smtp_user=email_cfg["smtp_user"],
        smtp_password=email_cfg["smtp_password"],
        mail_to=email_cfg["mail_to"],
        subject=subject,
        body="\n".join(body_lines),
        attachment=args.output,
    )
    print(f"邮件已发送至 {email_cfg['mail_to']}")


if __name__ == "__main__":
    main()
